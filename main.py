from openpyxl import load_workbook, Workbook


def open_workbook():
    workbook = load_workbook('carteira.xlsx')
    return workbook.active


def create_workbook():
    workbook = Workbook()
    workbook.active.title = 'Transações'
    workbook.save('carteira.xlsx')


try:
    worksheet = open_workbook()
except FileNotFoundError:
    create_workbook()
    worksheet = open_workbook()
