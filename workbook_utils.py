from openpyxl import load_workbook, Workbook


def open_workbook(file_name):
    workbook = load_workbook(file_name)
    return workbook, workbook.active


def create_workbook():
    workbook = Workbook()
    workbook.active.title = 'Transações'
    workbook.active.append(['Data', 'Categoria', 'Ordem', 'Ticker', 'Quantidade', 'Preço (R$)', 'Preço Total (R$)'])
    workbook.save('carteira.xlsx')


def save_workbook(workbook):
    workbook.save('carteira.xlsx')


def list_rows(worksheet):
    rows = []
    for row in list(worksheet.rows)[1:]:
        rows.append([cell.value for cell in row])
    return rows
